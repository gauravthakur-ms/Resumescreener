"""Excel export utility — generates downloadable Excel from Cosmos candidate data."""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
import io


def generate_excel_report(candidates: list[dict], jd_title: str = "") -> bytes:
    """Generate an Excel report from candidate records.
    
    Returns the Excel file as bytes (ready for Blob upload).
    Includes two sheets: 'All Candidates' and 'Top 10 Ranking'.
    """
    wb = Workbook()

    # Sheet 1: All Candidates
    ws_all = wb.active
    ws_all.title = "All Candidates"

    headers = [
        "#", "Name", "Email", "Phone", "Location", "Current Org",
        "Total Exp (yrs)", "Relevant Exp (yrs)", "Match Score",
        "Recommendation", "Screening", "Domain", "Certifications",
        "Risk Flags", "Rejection Reasons", "Last Work Date",
        "Recruiter Summary", "Classification", "File Name"
    ]
    ws_all.append(headers)

    # Style headers
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col_idx, _ in enumerate(headers, 1):
        cell = ws_all.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    # Sort by match_score descending
    sorted_candidates = sorted(
        candidates,
        key=lambda c: c.get("scoring", {}).get("match_score", 0),
        reverse=True
    )

    for idx, c in enumerate(sorted_candidates, 1):
        personal = c.get("personal", {})
        experience = c.get("experience", {})
        scoring = c.get("scoring", {})
        row = [
            idx,
            personal.get("name", "NDATA"),
            personal.get("email", "NDATA"),
            personal.get("phone", "NDATA"),
            personal.get("location", "NDATA"),
            personal.get("current_organization", "NDATA"),
            experience.get("total_years", 0),
            experience.get("relevant_years", 0),
            scoring.get("match_score", 0),
            c.get("recommendation", ""),
            c.get("screening", ""),
            c.get("domain_classification", ""),
            ", ".join(c.get("certifications", [])),
            "; ".join(c.get("risk_flags", [])),
            "; ".join(c.get("rejection_reasons", [])),
            experience.get("last_work_date", "NDATA"),
            c.get("recruiter_summary", ""),
            c.get("classification_folder", ""),
            c.get("file_name", ""),
        ]
        ws_all.append(row)

    # Auto-width columns
    for col in ws_all.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_all.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # Sheet 2: Top 10 Ranking
    ws_top = wb.create_sheet("Top 10 Ranking")
    top_headers = [
        "Rank", "Name", "Match Score", "Recommendation",
        "Total Exp", "Relevant Exp", "Domain", "Recruiter Summary"
    ]
    ws_top.append(top_headers)

    for col_idx, _ in enumerate(top_headers, 1):
        cell = ws_top.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

    for idx, c in enumerate(sorted_candidates[:10], 1):
        personal = c.get("personal", {})
        experience = c.get("experience", {})
        scoring = c.get("scoring", {})
        row = [
            idx,
            personal.get("name", "NDATA"),
            scoring.get("match_score", 0),
            c.get("recommendation", ""),
            experience.get("total_years", 0),
            experience.get("relevant_years", 0),
            c.get("domain_classification", ""),
            c.get("recruiter_summary", ""),
        ]
        ws_top.append(row)

    for col in ws_top.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_top.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
